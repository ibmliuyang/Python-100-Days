import openpyxl

# 打开源文件
src_wb = openpyxl.load_workbook('/Users/ly/Downloads/2021-yanfajiaji-fuzhu1.xlsx')
# 创建目标文件
dst_wb = openpyxl.Workbook()
dst_sheet = dst_wb.active
dst_sheet.title = 'Summary'

# 写入标题
dst_sheet['A1'] = 'A2原值'
dst_sheet['B1'] = 'G2原值'
dst_sheet['C1'] = 'F列合计'

# 遍历所有sheet页
for src_sheet in src_wb:
    # 计算A2、G2的值
    a2_value = src_sheet['A2'].value
    g2_value = src_sheet['G2'].value

    # 计算F列的和
    f_sum = 0
    for row in src_sheet.iter_rows(min_row=2, max_col=6, values_only=True):
        try:
            if row[5] is not None:
                f_sum += float(row[5])
        except ValueError:
            pass

    # 写入结果
    dst_sheet.append([a2_value, g2_value, f_sum])

# 保存目标文件
dst_wb.save('/Users/ly/Downloads/2021-yanfajiaji-fuzhu1-new.xlsx')