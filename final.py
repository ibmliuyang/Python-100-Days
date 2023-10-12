import openpyxl

# 打开源文件
src_wb = openpyxl.load_workbook('/Users/ly/Downloads/2021-yanfajiaji-fuzhu1.xlsx')
# 创建目标文件
dst_wb = openpyxl.Workbook()
dst_sheet = dst_wb.active
dst_sheet.title = 'Summary'
##  使用python的openpyxl获取所有sheet页中A2、G2、F列的所有数字的合计，放到另一个excel的A、B、C列

# 写入标题
dst_sheet['A1'] = 'A2原值'
dst_sheet['B1'] = 'G2原值'
dst_sheet['C1'] = 'S2原值'
dst_sheet['D1'] = 'U2原值'
dst_sheet['E1'] = 'AE2原值'
dst_sheet['F1'] = 'AG2原值'
dst_sheet['G1'] = 'H3原值'
dst_sheet['H1'] = 'J3原值'
dst_sheet['I1'] = 'P3原值'
dst_sheet['J1'] = 'R3原值'

dst_sheet['K1'] = 'F列合计'

# 遍历所有sheet页
for src_sheet in src_wb:
    # 计算A2、G2的值
    a2_value = src_sheet['A2'].value
    g2_value = src_sheet['G2'].value
    s2_value = src_sheet['S2'].value
    u2_value = src_sheet['U2'].value
    ae2_value = src_sheet['AE2'].value
    ag2_value = src_sheet['AG2'].value
    h3_value = src_sheet['H3'].value
    j3_value = src_sheet['J3'].value
    p3_value = src_sheet['P3'].value
    r3_value = src_sheet['R3'].value

    # 计算F列的和
    f_sum = 0
    for row in src_sheet.iter_rows(min_row=2, max_col=6, values_only=True):
        try:
            if row[5] is not None:
                f_sum += float(row[5])
        except ValueError:
            pass

    # 写入结果
    dst_sheet.append([a2_value, g2_value,  s2_value, u2_value, ae2_value, ag2_value, h3_value, j3_value, p3_value, r3_value, f_sum])

# 保存目标文件
dst_wb.save('/Users/ly/Downloads/2021-yanfajiaji-fuzhu1-new-2.xlsx')