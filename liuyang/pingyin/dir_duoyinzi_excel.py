import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from xpinyin import Pinyin

# 设置目录路径
directory_path = "/Users/ly/Downloads/手工税种-建表excel"

# 初始化拼音转换器
p = Pinyin()

# 遍历指定目录中的Excel文件
for filename in os.listdir(directory_path):
    if filename.endswith(".xlsx"):  # 只处理Excel文件
        file_path = os.path.join(directory_path, filename)

        # 读取Excel文件
        df = pd.read_excel(file_path)

        # 创建新的Excel文件名
        new_filename = os.path.splitext(filename)[0] + "_pinyin.xlsx"
        new_file_path = os.path.join(directory_path, new_filename)

        # 打开新创建的Excel文件以进行格式设置
        workbook = load_workbook(new_file_path)
        worksheet = workbook.active

        # 定义一个函数，用于将多音字的拼音首字母标记为红色
        def format_multi_pinyin(text):
            rich_text = []
            for char in text:
                font = Font(color="FF0000") if len(p.get_pinyin(char)) > 1 else Font()
                fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色背景
                cell_value = char
                if font.color and fill.start_color:  # 如果是多音字，设置字体颜色和背景色
                    cell_value = char
                else:
                    cell_value = char
                cell = worksheet.cell(row=row_idx, column=2)
                cell.value = cell_value
                if font.color:
                    cell.font = font
                if fill.start_color:
                    cell.fill = fill

        # 设置Excel文件中第二列的值和样式
        for row_idx, cell_value in enumerate(df['拼音'], start=2):
            format_multi_pinyin(cell_value)

        # 设置Excel文件中的列宽以适应内容
        for row in dataframe_to_rows(df, index=False, header=False):
            max_length = max(len(str(value) if value is not None else '') for value in row)
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions['B'].width = adjusted_width

        # 保存设置后的Excel文件
        workbook.save(new_file_path)

        print(f"已处理文件: {filename}")

print("所有文件处理完成")
